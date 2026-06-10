#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Stroytorg price XLS parser + validator.

Usage:
  python tools/price_parser_validator.py \
    --xls ./Prays_List_Stroytorg_novy_09_06_2026_g.xlsx \
    --catalog ./price_catalog.json \
    --mapping ./config/price_mapping.json \
    --out ./price_catalog.generated.json \
    --report ./reports/price_validation_report.json

Exit codes:
  0 - validation passed, generated JSON is safe to review/publish
  1 - validation has blocking errors
"""
import argparse, copy, datetime as dt, json, math, re, sys
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

try:
    from openpyxl import load_workbook
except Exception as exc:
    print('ERROR: openpyxl is required. Install: pip install openpyxl', file=sys.stderr)
    raise

PRICE_KEYS = ['grey','mono_color1','mono_color2','colormix_2','colormix_3','colormix_fakt','stonemix']
PRICE_COLS = { 'grey':4, 'mono_color1':5, 'mono_color2':6, 'colormix_2':7, 'colormix_3':8, 'colormix_fakt':9, 'stonemix':10 }
PACK_COLS_PAVER = {'pcs_per_pallet':11,'pcs_per_m2':12,'m2_per_pallet':13,'pallet_weight_kg':14}
PACK_COLS_CURB = {'pcs_per_pallet':11,'pcs_per_meter':12,'meters_per_pallet':13,'pallet_weight_kg':14}


def norm_text(v: Any) -> str:
    s = '' if v is None else str(v)
    s = s.lower().replace('ё','е')
    s = s.replace('"',' ').replace('«',' ').replace('»',' ')
    s = re.sub(r'[^a-zа-я0-9\.]+',' ',s)
    return re.sub(r'\s+',' ',s).strip()


def parse_mm(v: Any) -> Optional[int]:
    if v is None: return None
    m = re.search(r'(\d+)\s*мм', str(v).lower())
    if m: return int(m.group(1))
    m = re.search(r'\b(\d{2,3})\b', str(v))
    return int(m.group(1)) if m else None


def val(v: Any) -> Any:
    if v is None: return None
    if isinstance(v, (int,float)) and not isinstance(v,bool):
        if math.isnan(v): return None
        return round(float(v), 2)
    s = str(v).strip()
    if not s: return None
    low = s.lower().replace('ё','е')
    if 'запрос' in low: return 'по запросу'
    if low in {'-','—','–'}: return '-'
    s2 = low.replace(' ', '').replace(',', '.')
    try: return round(float(s2), 2)
    except Exception: return s


def numeric(v: Any) -> Optional[float]:
    if isinstance(v,(int,float)) and not isinstance(v,bool): return float(v)
    return None


def read_xls_rows(xls_path: Path, sheet: Optional[str]=None) -> Tuple[List[Dict[str,Any]], Dict[str,Any]]:
    wb = load_workbook(xls_path, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows=[]; section=None
    effective_from=None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        a = row[0] if len(row)>0 else None
        nt = norm_text(a)
        full = ' '.join(norm_text(x) for x in row if x is not None)
        if 'камень бортовой' in full:
            section='curbstone'; continue
        if 'плитка тротуарная' in full:
            section='paver'; continue
        mdate = re.search(r'действует\s+с\s+(\d{2})\.(\d{2})\.(\d{4})', full)
        if mdate:
            effective_from=f"{mdate.group(3)}-{mdate.group(2)}-{mdate.group(1)}"
        # data rows have name and at least grey price column
        if not a or row[3] is None or section not in {'paver','curbstone'}: continue
        if any(x in nt for x in ['наименование','продукция','цвет 1','цены указаны']): continue
        rec={'row':i,'section':section,'raw_name':str(a),'norm_name':nt,'height_or_width_mm':parse_mm(row[1] if len(row)>1 else None)}
        rec['prices']={k: val(row[col-1] if len(row)>=col else None) for k,col in PRICE_COLS.items()}
        cols = PACK_COLS_PAVER if section=='paver' else PACK_COLS_CURB
        rec['packaging']={k: val(row[col-1] if len(row)>=col else None) for k,col in cols.items()}
        rows.append(rec)
    meta={'sheet':ws.title,'sheets':wb.sheetnames,'effective_from':effective_from}
    return rows, meta


def row_matches(item: Dict[str,Any], row: Dict[str,Any]) -> bool:
    mt=item.get('match',{})
    if item.get('type') != row.get('section'): return False
    aliases=[norm_text(a) for a in mt.get('aliases',[]) if a]
    if not aliases: aliases=[norm_text(item.get('name'))]
    if not any(a and re.search(r'(?<![a-zа-я0-9])' + re.escape(a) + r'(?![a-zа-я0-9])', row['norm_name']) for a in aliases): return False
    if item.get('type')=='paver':
        return int(mt.get('thickness_mm')) == int(row.get('height_or_width_mm') or -1)
    if item.get('type')=='curbstone':
        width = mt.get('width_mm')
        return width is None or int(width) == int(row.get('height_or_width_mm') or -1)
    return False


def find_matches(mapping_item: Dict[str,Any], rows: List[Dict[str,Any]]) -> List[Dict[str,Any]]:
    return [r for r in rows if row_matches(mapping_item, r)]



def ensure_logistics_fields(catalog: Dict[str,Any]) -> None:
    """Populate future logistics fields from packaging without affecting current price calculation."""
    meta = catalog.setdefault('meta', {})
    meta['logistics_schema_version'] = '1.0.0'
    meta['logistics_status'] = 'prepared_not_active'
    meta['logistics_note'] = 'Logistics fields are informational until admin_settings.logistics.enabled=true and separate logistics module is connected.'

    for item in catalog.get('pavers', {}).get('items', []):
        for th in item.get('thickness', []):
            pack = th.get('packaging', {}) or {}
            units = numeric(pack.get('m2_per_pallet'))
            pallet_weight = numeric(pack.get('pallet_weight_kg'))
            th['logistics'] = {
                'unit_for_logistics': 'm2',
                'units_per_pallet': units,
                'pallet_weight_kg': pallet_weight,
                'weight_per_unit_kg': round(pallet_weight / units, 4) if pallet_weight and units else None,
                'volume_per_unit_m3': th.get('logistics', {}).get('volume_per_unit_m3'),
                'load_group': 'paver_tile',
                'enabled_for_future_logistics': True
            }

    for item in catalog.get('curbstone', {}).get('items', []):
        pack = item.get('packaging', {}) or {}
        units = numeric(pack.get('meters_per_pallet'))
        pallet_weight = numeric(pack.get('pallet_weight_kg'))
        item['logistics'] = {
            'unit_for_logistics': 'meter',
            'units_per_pallet': units,
            'pcs_per_pallet': numeric(pack.get('pcs_per_pallet')),
            'pallet_weight_kg': pallet_weight,
            'weight_per_unit_kg': round(pallet_weight / units, 4) if pallet_weight and units else None,
            'volume_per_unit_m3': item.get('logistics', {}).get('volume_per_unit_m3'),
            'load_group': 'curbstone',
            'enabled_for_future_logistics': True
        }


def update_catalog(base: Dict[str,Any], mapping: Dict[str,Any], rows: List[Dict[str,Any]]) -> Tuple[Dict[str,Any], Dict[str,Any]]:
    new=copy.deepcopy(base)
    errors=[]; warnings=[]; changes=[]; unchanged=[]; matched_rows=[]
    now=dt.datetime.now(dt.timezone.utc).isoformat()

    # index catalog references
    pavers={(it['id'], th['mm']):(it,th) for it in new.get('pavers',{}).get('items',[]) for th in it.get('thickness',[])}
    curbs={it['id']:it for it in new.get('curbstone',{}).get('items',[])}

    for mi in mapping.get('items',[]):
        matches=find_matches(mi, rows)
        if len(matches)==0:
            (errors if mi.get('required',True) else warnings).append({'sku':mi.get('sku'),'type':'missing_match','message':'Позиция не найдена в XLS','item':mi})
            continue
        target_prices=None; target_packaging=None
        if mi['type']=='paver':
            ref=pavers.get((mi['catalog_id'], int(mi['match']['thickness_mm'])))
            if not ref:
                errors.append({'sku':mi.get('sku'),'type':'catalog_missing','message':'SKU есть в mapping, но отсутствует в price_catalog.json'})
                continue
            _, th = ref; target_prices=th['prices']; target_packaging=th.get('packaging',{})
        else:
            ref=curbs.get(mi['catalog_id'])
            if not ref:
                errors.append({'sku':mi.get('sku'),'type':'catalog_missing','message':'SKU есть в mapping, но отсутствует в price_catalog.json'})
                continue
            target_prices=ref['prices']; target_packaging=ref.get('packaging',{})

        if len(matches)>1:
            old_grey = target_prices.get('grey') if target_prices else None
            same_old = [m for m in matches if numeric(m.get('prices',{}).get('grey')) is not None and numeric(old_grey) is not None and abs(numeric(m['prices']['grey']) - numeric(old_grey)) < 0.01]
            same_prices = []
            first_prices = matches[0].get('prices', {})
            if all(m.get('prices', {}) == first_prices for m in matches):
                same_prices = matches
            if len(same_old) == 1:
                warnings.append({'sku':mi.get('sku'),'type':'duplicate_resolved_by_old_price','message':'Несколько строк XLS; выбрана строка, где текущая серая цена совпадает с каталогом','rows':[m['row'] for m in matches],'selected_row':same_old[0]['row']})
                matches = same_old
            elif same_prices:
                warnings.append({'sku':mi.get('sku'),'type':'duplicate_identical_prices','message':'Несколько строк XLS с одинаковыми ценами; выбрана первая','rows':[m['row'] for m in matches],'selected_row':matches[0]['row']})
                matches = [matches[0]]
            else:
                errors.append({'sku':mi.get('sku'),'type':'duplicate_match','message':'Позиция сопоставилась с несколькими строками XLS','rows':[m['row'] for m in matches]})
                continue
        r=matches[0]; matched_rows.append(r['row'])

        for k in PRICE_KEYS:
            old=target_prices.get(k); newv=r['prices'].get(k)
            if newv is None:
                errors.append({'sku':mi['sku'],'type':'empty_price','price_key':k,'row':r['row'],'message':'Пустая цена в XLS'})
                continue
            if numeric(newv)==0:
                errors.append({'sku':mi['sku'],'type':'zero_price','price_key':k,'row':r['row'],'message':'Нулевая цена в XLS'})
                continue
            if numeric(old) is not None and numeric(newv) is not None:
                diff_pct=(numeric(newv)-numeric(old))/numeric(old)*100 if numeric(old) else 0
                if abs(diff_pct)>=50:
                    warnings.append({'sku':mi['sku'],'type':'critical_price_delta','price_key':k,'old':old,'new':newv,'delta_pct':round(diff_pct,2),'row':r['row']})
                elif abs(diff_pct)>=30:
                    warnings.append({'sku':mi['sku'],'type':'large_price_delta','price_key':k,'old':old,'new':newv,'delta_pct':round(diff_pct,2),'row':r['row']})
            if old != newv:
                changes.append({'sku':mi['sku'],'price_key':k,'old':old,'new':newv,'row':r['row']})
                target_prices[k]=newv
            else:
                unchanged.append({'sku':mi['sku'],'price_key':k,'value':old,'row':r['row']})
        # packaging is updated where compatible keys exist; additional logistics fields are not touched
        for pk,pv in r.get('packaging',{}).items():
            if pv is not None:
                target_packaging[pk]=pv

    # rows in XLS that look like product rows but were not mapped
    extra=[{'row':r['row'],'section':r['section'],'name':r['raw_name']} for r in rows if r['row'] not in set(matched_rows)]
    if extra:
        warnings.append({'type':'unmapped_xls_rows','message':'В XLS есть строки, не сопоставленные с калькулятором','count':len(extra),'rows':extra[:30]})

    ensure_logistics_fields(new)
    new.setdefault('meta',{})['generated_at']=now
    new['meta']['source_file']='XLS upload'
    new['meta']['schema_updated_by']='tools/price_parser_validator.py'

    report={'status':'blocked' if errors else 'passed','summary':{'mapping_items':len(mapping.get('items',[])),'xls_product_rows':len(rows),'changed':len(changes),'unchanged':len(unchanged),'errors':len(errors),'warnings':len(warnings)},'errors':errors,'warnings':warnings,'changes':changes[:1000]}
    return new, report


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--xls', required=True)
    ap.add_argument('--catalog', default='price_catalog.json')
    ap.add_argument('--mapping', default='config/price_mapping.json')
    ap.add_argument('--out', default='price_catalog.generated.json')
    ap.add_argument('--report', default='reports/price_validation_report.json')
    ap.add_argument('--sheet', default=None)
    args=ap.parse_args()

    xls=Path(args.xls); catalog_p=Path(args.catalog); mapping_p=Path(args.mapping)
    catalog=json.load(open(catalog_p,encoding='utf-8'))
    mapping=json.load(open(mapping_p,encoding='utf-8'))
    rows, xls_meta=read_xls_rows(xls, args.sheet)
    new, report=update_catalog(catalog, mapping, rows)
    if xls_meta.get('effective_from'):
        new.setdefault('meta',{})['effective_from']=xls_meta['effective_from']
    new.setdefault('meta',{})['source_xls_sheet']=xls_meta.get('sheet')
    new.setdefault('meta',{})['source_file']=xls.name
    report['xls_meta']=xls_meta

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    Path(args.report).parent.mkdir(parents=True, exist_ok=True)
    json.dump(new, open(args.out,'w',encoding='utf-8'), ensure_ascii=False, indent=2)
    json.dump(report, open(args.report,'w',encoding='utf-8'), ensure_ascii=False, indent=2)
    print(json.dumps({'status':report['status'], **report['summary'], 'out':args.out, 'report':args.report}, ensure_ascii=False, indent=2))
    return 1 if report['errors'] else 0

if __name__=='__main__':
    raise SystemExit(main())
