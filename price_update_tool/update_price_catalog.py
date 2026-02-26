#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Update repo_bundle/price_catalog.json from supplier XLSX (Актив Групп).
Designed for repo: miropolcevpro/prev_img_ag_2_026 (structure matches current price_catalog.json).

Usage:
  python3 update_price_catalog.py --xlsx "Прайс.xlsx" --price-catalog "./price_catalog.json" --out "./price_catalog.json" --report "./report.json"

Notes:
- We UPDATE existing JSON structure (ids preserved). We match items by NAME and thickness mm.
- If some rows cannot be mapped to existing items, script fails (non-zero exit) and writes report.
"""
import argparse, json, re, sys
from pathlib import Path
import openpyxl

PRICE_KEYS = [
  ("grey", 4),
  ("mono_color1", 5),
  ("mono_color2", 6),
  ("colormix_2", 7),
  ("colormix_3", 8),
  ("colormix_fakt", 9),
  ("stonemix", 10),
]

def norm_str(s):
  if s is None: return None
  s = str(s).strip()
  s = re.sub(r"\s+", " ", s)
  return s

def to_value(v):
  # keep numbers as float/int, keep strings as trimmed
  if v is None: return None
  if isinstance(v, (int, float)):
    return float(v) if isinstance(v, float) else int(v)
  return str(v).strip()

def extract_form_name(cell_text):
  if not cell_text: return None
  m = re.search(r'"([^"]+)"', str(cell_text))
  if m: return norm_str(m.group(1))
  # curbstone rows don't have quotes; use leading token like "БР 100.20.8"
  # but we match curbstone by full name later, so return full cleaned text
  return None

def build_name_index(price_catalog):
  idx = {}
  for it in price_catalog.get("pavers", {}).get("items", []):
    n = norm_str(it.get("name"))
    if n:
      idx[n.lower()] = it
  curb_idx = {}
  for it in price_catalog.get("curbstone", {}).get("items", []):
    n = norm_str(it.get("name"))
    if n:
      curb_idx[n.lower()] = it
  return idx, curb_idx

def update_pavers(ws, price_catalog, report):
  pavers_idx, _ = build_name_index(price_catalog)
  updated = 0
  missing = []
  # rows 7.. until we hit curbstone header "КАМЕНЬ БОРТОВОЙ"
  max_row = ws.max_row
  for r in range(7, max_row+1):
    a = ws.cell(r,1).value
    if isinstance(a,str) and "КАМЕНЬ БОРТОВОЙ" in a.upper():
      break
    h = ws.cell(r,2).value
    if not a or not h: 
      continue
    form_name = extract_form_name(a)
    if not form_name:
      continue
    form = pavers_idx.get(form_name.lower())
    if not form:
      missing.append({"row": r, "name": form_name})
      continue
    # thickness mm from col2 like "40 мм"
    mm_m = re.search(r"(\d+)", str(h))
    if not mm_m:
      missing.append({"row": r, "name": form_name, "reason":"no thickness"})
      continue
    mm = int(mm_m.group(1))
    # find thickness entry
    th_list = form.get("thickness", [])
    th = None
    for t in th_list:
      if int(t.get("mm", -1)) == mm:
        th = t
        break
    if th is None:
      # add new thickness entry to preserve future; but report it
      th = {"mm": mm, "prices": {}, "packaging": {}}
      th_list.append(th)
      form["thickness"] = th_list
      report["added_thickness"].append({"name": form_name, "mm": mm, "row": r})
    # prices
    prices = th.setdefault("prices", {})
    for key, col in PRICE_KEYS:
      prices[key] = to_value(ws.cell(r,col).value)
    # packaging
    pack = th.setdefault("packaging", {})
    pack["pcs_per_pallet"] = to_value(ws.cell(r,11).value)
    pack["pcs_per_m2"] = to_value(ws.cell(r,12).value)
    pack["m2_per_pallet"] = to_value(ws.cell(r,13).value)
    pack["pallet_weight_kg"] = to_value(ws.cell(r,14).value)
    updated += 1
  report["pavers_rows_updated"] = updated
  report["pavers_missing"] = missing
  return len(missing) == 0

def update_curbstone(ws, price_catalog, report):
  _, curb_idx = build_name_index(price_catalog)
  updated = 0
  missing = []
  # curbstone starts at row where col1 contains "БР "
  for r in range(34, ws.max_row+1):
    a = ws.cell(r,1).value
    if not a or not isinstance(a,str): 
      continue
    if a.strip().startswith("БР"):
      # build clean name same as in json (it includes size at end)
      # In sheet it's like 'БР 100.20.8 1000х200х80'
      name = norm_str(a)
      # Try exact match
      item = curb_idx.get(name.lower())
      if not item:
        # sometimes multiple spaces; try normalize: remove double spaces
        name2 = re.sub(r"\s+", " ", name)
        item = curb_idx.get(name2.lower())
      if not item:
        missing.append({"row": r, "name": name})
        continue
      prices = item.setdefault("prices", {})
      for key, col in PRICE_KEYS:
        prices[key] = to_value(ws.cell(r,col).value)
      pack = item.setdefault("packaging", {})
      pack["pcs_per_pallet"] = to_value(ws.cell(r,11).value)
      pack["pallet_weight_kg"] = to_value(ws.cell(r,14).value)
      # keep extra keys if any; optionally add lm data safely
      pack.setdefault("pcs_per_lm", to_value(ws.cell(r,12).value))
      pack.setdefault("lm_per_pallet", to_value(ws.cell(r,13).value))
      updated += 1
  report["curb_rows_updated"] = updated
  report["curb_missing"] = missing
  return len(missing) == 0

def main():
  ap = argparse.ArgumentParser()
  ap.add_argument("--xlsx", required=True)
  ap.add_argument("--price-catalog", required=True, help="Path to existing price_catalog.json")
  ap.add_argument("--out", default=None, help="Output path (default overwrite --price-catalog)")
  ap.add_argument("--sheet", default="плитка+борт")
  ap.add_argument("--report", default="price_update_report.json")
  args = ap.parse_args()

  xlsx_path = Path(args.xlsx)
  pc_path = Path(args.price_catalog)
  out_path = Path(args.out) if args.out else pc_path
  report_path = Path(args.report)

  report = {"xlsx": str(xlsx_path), "sheet": args.sheet, "pavers_rows_updated": 0, "curb_rows_updated": 0,
            "pavers_missing": [], "curb_missing": [], "added_thickness": []}

  if not xlsx_path.exists():
    print("XLSX not found:", xlsx_path, file=sys.stderr); sys.exit(2)
  if not pc_path.exists():
    print("price_catalog.json not found:", pc_path, file=sys.stderr); sys.exit(2)

  price_catalog = json.loads(pc_path.read_text(encoding="utf-8"))
  wb = openpyxl.load_workbook(xlsx_path, data_only=True)
  if args.sheet not in wb.sheetnames:
    print("Sheet not found. Available:", wb.sheetnames, file=sys.stderr); sys.exit(2)
  ws = wb[args.sheet]

  ok1 = update_pavers(ws, price_catalog, report)
  ok2 = update_curbstone(ws, price_catalog, report)

  report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

  if not (ok1 and ok2):
    print("Update failed: some items could not be mapped. See report:", report_path, file=sys.stderr)
    sys.exit(3)

  out_path.write_text(json.dumps(price_catalog, ensure_ascii=False), encoding="utf-8")
  print("OK: updated", out_path)
  print("Report:", report_path)

if __name__ == "__main__":
  main()
